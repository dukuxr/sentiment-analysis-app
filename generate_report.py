"""
Generate Excel Report with Formulas
All calculations are done with Excel formulas so you can edit values and see updates automatically
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

def create_excel_with_formulas(csv_file='algorithm_comparison_results.csv'):
    """
    Create Excel report with formulas for dynamic calculations
    """
    
    print("="*70)
    print("GENERATING EXCEL REPORT WITH FORMULAS")
    print("="*70)
    
    # Check if CSV exists
    if not os.path.exists(csv_file):
        print(f"\n[ERROR] Cannot find {csv_file}")
        print("Please run the comparison analysis first!")
        return
    
    # Load results
    df_results = pd.read_csv(csv_file)
    
    # Create Excel writer
    output_file = 'Thesis_Report_With_Formulas.xlsx'
    
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        
        # Write raw data first
        df_results.to_excel(writer, sheet_name='Raw Data', index=False)
    
    # Now add formulas
    wb = load_workbook(output_file)
    
    # ========================================================================
    # SHEET 1: EXECUTIVE SUMMARY WITH FORMULAS
    # ========================================================================
    
    ws_summary = wb.create_sheet('Executive Summary', 0)
    
    # Headers
    ws_summary['A1'] = 'Item'
    ws_summary['B1'] = 'Value'
    
    # Data with formulas
    summary_items = [
        ('Project Title', 'Comparative Analysis of ML Algorithms for Sentiment Classification'),
        ('Student Name', '[Your Name Here]'),
        ('Date', '=TODAY()'),  # Formula: Today's date
        ('Dataset', 'IMDB Movie Reviews'),
        ('Total Samples', '25000'),
        ('Training Samples', '20000'),
        ('Testing Samples', '5000'),
        ('Algorithms Tested', '=COUNTA(\'Raw Data\'!A2:A100)'),  # Formula: Count algorithms
        ('Best Algorithm', '=INDEX(\'Raw Data\'!A:A,MATCH(MAX(\'Raw Data\'!B:B),\'Raw Data\'!B:B,0))'),  # Formula: Algorithm with max accuracy
        ('Best Accuracy (%)', '=MAX(\'Raw Data\'!B:B)'),  # Formula: Max accuracy
        ('Status', 'Experimental Phase Complete')
    ]
    
    for idx, (item, value) in enumerate(summary_items, start=2):
        ws_summary[f'A{idx}'] = item
        if isinstance(value, str) and value.startswith('='):
            ws_summary[f'B{idx}'] = value  # Formula
        else:
            ws_summary[f'B{idx}'] = value  # Static value
    
    # ========================================================================
    # SHEET 2: STATISTICS WITH FORMULAS
    # ========================================================================
    
    ws_stats = wb.create_sheet('Statistics')
    
    # Headers
    headers = ['Metric', 'Mean', 'Std Dev', 'Min', 'Max', 'Range']
    for col, header in enumerate(headers, start=1):
        ws_stats.cell(1, col, header)
    
    # Metrics (rows)
    metrics = [
        ('Accuracy (%)', 'B', 2),      # Column B in Raw Data, starting row 2
        ('Precision (%)', 'C', 2),
        ('Recall (%)', 'D', 2),
        ('F1-Score (%)', 'E', 2),
        ('Training Time (s)', 'F', 2),
        ('Prediction Time (s)', 'G', 2)
    ]
    
    for row_idx, (metric_name, col_letter, start_row) in enumerate(metrics, start=2):
        ws_stats[f'A{row_idx}'] = metric_name
        
        # Calculate last row dynamically
        num_algos = len(df_results)
        end_row = start_row + num_algos - 1
        
        # Formulas for statistics
        ws_stats[f'B{row_idx}'] = f"=AVERAGE('Raw Data'!{col_letter}{start_row}:{col_letter}{end_row})"  # Mean
        ws_stats[f'C{row_idx}'] = f"=STDEV('Raw Data'!{col_letter}{start_row}:{col_letter}{end_row})"    # Std Dev
        ws_stats[f'D{row_idx}'] = f"=MIN('Raw Data'!{col_letter}{start_row}:{col_letter}{end_row})"      # Min
        ws_stats[f'E{row_idx}'] = f"=MAX('Raw Data'!{col_letter}{start_row}:{col_letter}{end_row})"      # Max
        ws_stats[f'F{row_idx}'] = f"=E{row_idx}-D{row_idx}"                                               # Range
    
    # ========================================================================
    # SHEET 3: RANKINGS WITH FORMULAS
    # ========================================================================
    
    ws_rankings = wb.create_sheet('Rankings')
    
    # Headers
    ws_rankings['A1'] = 'Rank'
    ws_rankings['B1'] = 'Algorithm'
    ws_rankings['C1'] = 'Accuracy (%)'
    ws_rankings['D1'] = 'Precision (%)'
    ws_rankings['E1'] = 'Recall (%)'
    ws_rankings['F1'] = 'F1-Score (%)'
    
    # Ranking formulas (showing top performers)
    for rank in range(1, 6):  # Top 5
        row = rank + 1
        ws_rankings[f'A{row}'] = rank
        
        # LARGE function finds the Nth largest value
        # INDEX/MATCH to find corresponding algorithm name
        ws_rankings[f'B{row}'] = f"=IFERROR(INDEX('Raw Data'!A:A,MATCH(LARGE('Raw Data'!B:B,{rank}),'Raw Data'!B:B,0)),\"\")"
        ws_rankings[f'C{row}'] = f"=IFERROR(LARGE('Raw Data'!B:B,{rank}),\"\")"
        ws_rankings[f'D{row}'] = f"=IFERROR(INDEX('Raw Data'!C:C,MATCH(C{row},'Raw Data'!B:B,0)),\"\")"
        ws_rankings[f'E{row}'] = f"=IFERROR(INDEX('Raw Data'!D:D,MATCH(C{row},'Raw Data'!B:B,0)),\"\")"
        ws_rankings[f'F{row}'] = f"=IFERROR(INDEX('Raw Data'!E:E,MATCH(C{row},'Raw Data'!B:B,0)),\"\")"
    
    # ========================================================================
    # SHEET 4: KEY FINDINGS WITH FORMULAS
    # ========================================================================
    
    ws_findings = wb.create_sheet('Key Findings')
    
    # Headers
    ws_findings['A1'] = 'Category'
    ws_findings['B1'] = 'Algorithm'
    ws_findings['C1'] = 'Value'
    
    # Findings with formulas
    findings = [
        ('Best Accuracy', 
         "=INDEX('Raw Data'!A:A,MATCH(MAX('Raw Data'!B:B),'Raw Data'!B:B,0))",
         "=MAX('Raw Data'!B:B)&\"%\""),
        
        ('Best Precision', 
         "=INDEX('Raw Data'!A:A,MATCH(MAX('Raw Data'!C:C),'Raw Data'!C:C,0))",
         "=MAX('Raw Data'!C:C)&\"%\""),
        
        ('Best Recall', 
         "=INDEX('Raw Data'!A:A,MATCH(MAX('Raw Data'!D:D),'Raw Data'!D:D,0))",
         "=MAX('Raw Data'!D:D)&\"%\""),
        
        ('Best F1-Score', 
         "=INDEX('Raw Data'!A:A,MATCH(MAX('Raw Data'!E:E),'Raw Data'!E:E,0))",
         "=MAX('Raw Data'!E:E)&\"%\""),
        
        ('Fastest Training', 
         "=INDEX('Raw Data'!A:A,MATCH(MIN('Raw Data'!F:F),'Raw Data'!F:F,0))",
         "=MIN('Raw Data'!F:F)&\" sec\""),
        
        ('Fastest Prediction', 
         "=INDEX('Raw Data'!A:A,MATCH(MIN('Raw Data'!G:G),'Raw Data'!G:G,0))",
         "=MIN('Raw Data'!G:G)&\" sec\""),
        
        ('Average Accuracy', 
         "All Algorithms",
         "=AVERAGE('Raw Data'!B:B)&\"%\""),
    ]
    
    for idx, (category, algo_formula, value_formula) in enumerate(findings, start=2):
        ws_findings[f'A{idx}'] = category
        ws_findings[f'B{idx}'] = algo_formula
        ws_findings[f'C{idx}'] = value_formula
    
    # ========================================================================
    # SHEET 5: COMPARISON CHART DATA (for easy charting)
    # ========================================================================
    
    ws_chart = wb.create_sheet('Chart Data')
    
    ws_chart['A1'] = 'Algorithm'
    ws_chart['B1'] = 'Accuracy'
    ws_chart['C1'] = 'Precision'
    ws_chart['D1'] = 'Recall'
    ws_chart['E1'] = 'F1-Score'
    
    # Link to Raw Data
    for i in range(2, 2 + len(df_results)):
        ws_chart[f'A{i}'] = f"='Raw Data'!A{i}"
        ws_chart[f'B{i}'] = f"='Raw Data'!B{i}"
        ws_chart[f'C{i}'] = f"='Raw Data'!C{i}"
        ws_chart[f'D{i}'] = f"='Raw Data'!D{i}"
        ws_chart[f'E{i}'] = f"='Raw Data'!E{i}"
    
    # ========================================================================
    # SHEET 6: ABSTRACT FILLER (automatically fills abstract template)
    # ========================================================================
    
    ws_abstract = wb.create_sheet('Abstract Helper')
    
    ws_abstract['A1'] = 'Field'
    ws_abstract['B1'] = 'Value (Copy this for abstract)'
    
    abstract_fields = [
        ('Best Algorithm', "=INDEX('Raw Data'!A:A,MATCH(MAX('Raw Data'!B:B),'Raw Data'!B:B,0))"),
        ('Best Accuracy', "=ROUND(MAX('Raw Data'!B:B),2)"),
        ('Best Precision', "=ROUND(MAX('Raw Data'!C:C),2)"),
        ('Best Recall', "=ROUND(MAX('Raw Data'!D:D),2)"),
        ('Best F1-Score', "=ROUND(MAX('Raw Data'!E:E),2)"),
        ('Fastest Algorithm', "=INDEX('Raw Data'!A:A,MATCH(MIN('Raw Data'!F:F),'Raw Data'!F:F,0))"),
        ('Fastest Training Time', "=ROUND(MIN('Raw Data'!F:F),2)"),
        ('Slowest Training Time', "=ROUND(MAX('Raw Data'!F:F),2)"),
        ('Average Accuracy', "=ROUND(AVERAGE('Raw Data'!B:B),2)"),
        ('Number of Algorithms', "=COUNTA('Raw Data'!A2:A100)"),
    ]
    
    for idx, (field, formula) in enumerate(abstract_fields, start=2):
        ws_abstract[f'A{idx}'] = field
        ws_abstract[f'B{idx}'] = formula
    
    # ========================================================================
    # FORMAT ALL SHEETS
    # ========================================================================
    
    # Define styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Format headers
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 3, 60)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Format data rows
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = border
                if cell.column > 1:  # Numbers columns
                    cell.alignment = Alignment(horizontal='center')
                    # Format numbers to 2 decimal places
                    if cell.value and isinstance(cell.value, (int, float)):
                        cell.number_format = '0.00'
    
    # Save workbook
    wb.save(output_file)
    
    print(f"\n[SUCCESS] Created: {output_file}")
    print("\n" + "="*70)
    print("EXCEL FILE STRUCTURE:")
    print("="*70)
    print("\n1. Raw Data")
    print("   - Original data from CSV")
    print("   - Edit values here to see automatic updates everywhere")
    
    print("\n2. Executive Summary")
    print("   - Auto-calculated summary")
    print("   - Shows best algorithm and key stats")
    
    print("\n3. Statistics")
    print("   - Mean, Std Dev, Min, Max, Range")
    print("   - All calculated with formulas")
    
    print("\n4. Rankings")
    print("   - Top 5 algorithms by accuracy")
    print("   - Automatically sorted")
    
    print("\n5. Key Findings")
    print("   - Best performers in each category")
    print("   - Auto-updates when data changes")
    
    print("\n6. Chart Data")
    print("   - Ready for creating Excel charts")
    print("   - Select this data → Insert → Chart")
    
    print("\n7. Abstract Helper")
    print("   - Numbers ready to copy into your abstract")
    print("   - Copy column B values directly")
    
    print("\n" + "="*70)
    print("HOW TO USE:")
    print("="*70)
    print("\n✓ Edit numbers in 'Raw Data' sheet")
    print("✓ All other sheets update automatically")
    print("✓ Copy values from 'Abstract Helper' for congress submission")
    print("✓ Use 'Chart Data' to create graphs in Excel")
    print("✓ All formulas are visible - click any cell to see them")
    
    print("\n" + "="*70)
    print("CREATING CHARTS IN EXCEL:")
    print("="*70)
    print("\n1. Go to 'Chart Data' sheet")
    print("2. Select all data (A1:E6)")
    print("3. Click Insert → Recommended Charts")
    print("4. Choose Column Chart or Bar Chart")
    print("5. Format and save as image")
    
    return output_file


def create_abstract_template():
    """Create an abstract template with placeholders"""
    
    template = """
================================================================================
CONGRESS ABSTRACT TEMPLATE
Fill in values from 'Abstract Helper' sheet in Excel
================================================================================

TITLE:
Comparative Analysis of Machine Learning Algorithms for Sentiment 
Classification of IMDB Movie Reviews

ABSTRACT:
This study presents a comprehensive comparative analysis of [NUMBER] machine 
learning algorithms for binary sentiment classification of movie reviews. Using 
the IMDB dataset (25,000 reviews), we implemented and evaluated Logistic 
Regression, Naive Bayes, Support Vector Machine, Random Forest, and Decision 
Tree algorithms. 

Text preprocessing included HTML removal, lowercasing, and stopword filtering, 
with TF-IDF vectorization for feature extraction (5,000 features). The dataset 
was split 80-20 for training and testing. 

Results showed all algorithms exceeded 75% accuracy, with [BEST_ALGORITHM] 
achieving [BEST_ACCURACY]% accuracy, demonstrating superior precision 
([BEST_PRECISION]%), recall ([BEST_RECALL]%), and F1-score ([BEST_F1]%). 

Training time analysis revealed [FASTEST_ALGORITHM] as the fastest 
([FASTEST_TIME]s) while [BEST_ALGORITHM] provided optimal accuracy despite 
longer training time ([SLOWEST_TIME]s). 

The findings indicate that [BEST_ALGORITHM] offers optimal balance between 
accuracy and computational efficiency for practical sentiment analysis 
deployment. Future work includes exploring deep learning architectures 
(LSTM, BERT) and multi-class sentiment classification.

KEYWORDS: 
Sentiment Analysis, Machine Learning, Text Classification, Natural Language 
Processing, IMDB Dataset, Comparative Study

================================================================================
INSTRUCTIONS:
1. Open 'Thesis_Report_With_Formulas.xlsx'
2. Go to 'Abstract Helper' sheet
3. Copy values from column B
4. Replace [PLACEHOLDERS] above with those values
5. Proofread and submit!
================================================================================
    """
    
    with open('Abstract_Template.txt', 'w') as f:
        f.write(template)
    
    print("\n[SUCCESS] Created: Abstract_Template.txt")


if __name__ == "__main__":
    # Check for required package
    try:
        import openpyxl
    except ImportError:
        print("\n[ERROR] openpyxl package not found!")
        print("Install it with: pip install openpyxl")
        exit()
    
    # Generate Excel with formulas
    excel_file = create_excel_with_formulas()
    
    # Create abstract template
    create_abstract_template()
    
    print("\n" + "="*70)
    print("ALL FILES CREATED!")
    print("="*70)
    print(f"\n1. {excel_file}")
    print("2. Abstract_Template.txt")
    print("\nYou're ready to fill in your congress abstract! 🎉")