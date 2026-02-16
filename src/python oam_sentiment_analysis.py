"""
Export OAM Matrix to Excel for COCO System Analysis
Formats data exactly as needed for https://miau.my-x.hu/myx-free/coco/beker_std.php
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import re
from pathlib import Path

# Project paths
BASE_DIR = Path(__file__).resolve().parent.parent
DATA_DIR = BASE_DIR / "data"
REPORTS_DIR = BASE_DIR / "reports"

class COCOExporter:
    """
    Exports sentiment analysis data in COCO-compatible format
    """
    
    def __init__(self):
        self.positive_words = [
            'excellent', 'amazing', 'wonderful', 'brilliant', 'fantastic', 'great',
            'superb', 'outstanding', 'perfect', 'loved', 'best', 'beautiful',
            'awesome', 'incredible', 'masterpiece', 'recommend', 'enjoyed'
        ]
        
        self.negative_words = [
            'terrible', 'awful', 'horrible', 'worst', 'bad', 'boring', 'waste',
            'disappointing', 'poor', 'hate', 'avoid', 'garbage', 'dull',
            'pathetic', 'ridiculous', 'pointless', 'annoying'
        ]
    
    def extract_attributes(self, text):
        """Extract sentiment attributes from review text"""
        text = text.lower()
        
        attributes = {
            'positive_words': sum(1 for word in self.positive_words if word in text),
            'negative_words': sum(1 for word in self.negative_words if word in text),
            'exclamations': text.count('!'),
            'review_length': len(text.split()),
            'capital_ratio': sum(1 for c in text if c.isupper()) / max(len(text), 1) * 100,
            'questions': text.count('?'),
            'pos_neg_ratio': sum(1 for word in self.positive_words if word in text) / 
                            max(sum(1 for word in self.negative_words if word in text), 1),
            'word_diversity': len(set(text.split())) / max(len(text.split()), 1)
        }
        
        return attributes
    
    def create_oam_for_coco(self, csv_file=None, n_samples=50, output_file=None):
        """
        Create OAM matrix and export in COCO format
        """
        
        print("="*70)
        print("EXPORTING OAM MATRIX FOR COCO ANALYSIS")
        print("="*70)
        
        # Load dataset
        if csv_file is None:
            csv_file = DATA_DIR / "IMBD Dataset.csv"
        else:
            csv_file = Path(csv_file)

        if output_file is None:
            REPORTS_DIR.mkdir(parents=True, exist_ok=True)
            output_file = REPORTS_DIR / "OAM_for_COCO.xlsx"
        else:
            output_file = Path(output_file)

        print(f"\n1. Loading dataset from {csv_file}...")
        try:
            df = pd.read_csv(csv_file)
            print(f"   Total reviews: {len(df):,}")
        except Exception as e:
            print(f"   ERROR: {e}")
            return
        
        # Sample reviews (mix of positive and negative)
        print(f"\n2. Selecting {n_samples} reviews...")
        df['sentiment'] = df['sentiment'].str.lower()
        
        # Get balanced sample
        n_per_class = n_samples // 2
        positive_samples = df[df['sentiment'] == 'positive'].sample(n=n_per_class, random_state=42)
        negative_samples = df[df['sentiment'] == 'negative'].sample(n=n_per_class, random_state=42)
        
        sample_df = pd.concat([positive_samples, negative_samples]).sample(frac=1, random_state=42)
        print(f"   Selected: {len(sample_df)} reviews")
        print(f"   Positive: {len(positive_samples)}, Negative: {len(negative_samples)}")
        
        # Extract attributes for each review
        print(f"\n3. Extracting attributes...")
        oam_data = []
        
        for idx, row in sample_df.iterrows():
            review_text = row['review']
            sentiment = row['sentiment']
            
            # Get attributes
            attrs = self.extract_attributes(review_text)
            
            # Add metadata
            attrs['review_id'] = f"Review_{idx}"
            attrs['actual_sentiment'] = 1 if sentiment == 'positive' else 0
            attrs['review_preview'] = review_text[:100] + "..."  # First 100 chars
            
            oam_data.append(attrs)
        
        oam_df = pd.DataFrame(oam_data)
        print(f"   Extracted {len(oam_df.columns) - 3} attributes")  # Exclude ID, sentiment, preview
        
        # Reorder columns: metadata first, then attributes, then Y value (sentiment)
        columns_order = ['review_id', 'review_preview', 
                        'positive_words', 'negative_words', 'exclamations', 
                        'review_length', 'capital_ratio', 'questions', 
                        'pos_neg_ratio', 'word_diversity',
                        'actual_sentiment']
        
        oam_df = oam_df[columns_order]
        
        # Create Excel workbook
        print(f"\n4. Creating Excel file: {output_file}...")
        
        wb = Workbook()
        
        # ==================================================================
        # SHEET 1: OAM Matrix (Raw Data)
        # ==================================================================
        ws_oam = wb.active
        ws_oam.title = "OAM Matrix"
        
        # Write headers
        for col_idx, col_name in enumerate(oam_df.columns, start=1):
            cell = ws_oam.cell(1, col_idx, col_name)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Write data
        for row_idx, row in enumerate(oam_df.itertuples(index=False), start=2):
            for col_idx, value in enumerate(row, start=1):
                ws_oam.cell(row_idx, col_idx, value)
        
        # Auto-adjust column widths
        for column in ws_oam.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_oam.column_dimensions[column_letter].width = adjusted_width
        
        # ==================================================================
        # SHEET 2: COCO Input Format (Ranked Values)
        # ==================================================================
        ws_coco = wb.create_sheet("COCO Input Format")
        
        # Prepare data for COCO (exclude review_preview, sort by correlation)
        coco_df = oam_df.drop(['review_preview'], axis=1).copy()
        
        # Calculate correlations with sentiment
        attribute_cols = ['positive_words', 'negative_words', 'exclamations', 
                         'review_length', 'capital_ratio', 'questions', 
                         'pos_neg_ratio', 'word_diversity']
        
        correlations = {}
        for col in attribute_cols:
            correlations[col] = coco_df[col].corr(coco_df['actual_sentiment'])
        
        # Sort attributes by absolute correlation (strongest first)
        sorted_attrs = sorted(correlations.items(), key=lambda x: abs(x[1]), reverse=True)
        
        # Create sorted column order
        sorted_columns = ['review_id'] + [attr for attr, _ in sorted_attrs] + ['actual_sentiment']
        coco_sorted = coco_df[sorted_columns]
        
        # Rank each attribute (COCO needs ranked values)
        ranked_df = coco_sorted.copy()
        for col in attribute_cols:
            # Direction: 0 if positive correlation, 1 if negative
            direction = 0 if correlations[col] > 0 else 1
            
            if direction == 0:
                # Higher is better - rank ascending
                ranked_df[col] = coco_sorted[col].rank(method='dense').astype(int)
            else:
                # Lower is better - rank descending
                ranked_df[col] = coco_sorted[col].rank(method='dense', ascending=False).astype(int)
        
        # Multiply Y value (sentiment) by 1000 (COCO convention)
        ranked_df['actual_sentiment'] = ranked_df['actual_sentiment'] * 1000
        
        # Add metadata rows at top
        ws_coco['A1'] = "ATTRIBUTE TYPE"
        ws_coco['A2'] = "DIRECTION"
        ws_coco['A3'] = "CORRELATION"
        ws_coco['A4'] = "UNIT"
        
        # Fill metadata for each attribute
        col_start = 2  # Column B
        for idx, (attr, corr) in enumerate(sorted_attrs, start=col_start):
            # Row 1: Type (X for attributes, Y for target)
            ws_coco.cell(1, idx, "X")
            
            # Row 2: Direction (0 = positive, 1 = negative)
            direction = 0 if corr > 0 else 1
            ws_coco.cell(2, idx, direction)
            
            # Row 3: Correlation value
            ws_coco.cell(3, idx, round(corr, 4))
            
            # Row 4: Unit
            ws_coco.cell(4, idx, "count/ratio")
        
        # Y column (actual_sentiment)
        y_col = len(sorted_attrs) + col_start
        ws_coco.cell(1, y_col, "Y")
        ws_coco.cell(2, y_col, 0)  # Direction always 0 for Y
        ws_coco.cell(3, y_col, 1.0)  # Perfect correlation with itself
        ws_coco.cell(4, y_col, "*1000")
        
        # Row 5: Column headers
        ws_coco.cell(5, 1, "Object_ID")
        for idx, (attr, _) in enumerate(sorted_attrs, start=col_start):
            ws_coco.cell(5, idx, attr)
        ws_coco.cell(5, y_col, "Sentiment_Score")
        
        # Write ranked data starting from row 6
        for row_idx, row in enumerate(ranked_df.itertuples(index=False), start=6):
            for col_idx, value in enumerate(row, start=1):
                ws_coco.cell(row_idx, col_idx, value)
        
        # Format headers
        for cell in ws_coco[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        for row in [2, 3, 4, 5]:
            for cell in ws_coco[row]:
                if cell.value is not None:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center')
        
        # Auto-adjust columns
        for column in ws_coco.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws_coco.column_dimensions[column_letter].width = adjusted_width
        
        # ==================================================================
        # SHEET 3: Instructions
        # ==================================================================
        ws_instructions = wb.create_sheet("Instructions")
        
        instructions = [
            ["HOW TO USE WITH COCO SYSTEM"],
            [""],
            ["Step 1: Open the COCO Standard Interface"],
            ["   Go to: https://miau.my-x.hu/myx-free/coco/beker_std.php"],
            [""],
            ["Step 2: Prepare Your Data"],
            ["   - Go to 'COCO Input Format' sheet in this Excel file"],
            ["   - Copy the RANKED VALUES (starting from row 6, column B)"],
            ["   - Include ALL columns including the sentiment score"],
            [""],
            ["Step 3: Paste into COCO"],
            ["   - In the COCO interface, paste your data into the text boxes"],
            ["   - Each row = one object (review)"],
            ["   - Each column = one attribute"],
            ["   - Last column = Y value (sentiment * 1000)"],
            [""],
            ["Step 4: Run Analysis"],
            ["   - Click 'Calculate' or 'Run' button"],
            ["   - COCO will identify patterns and predict sentiment"],
            [""],
            ["Step 5: Interpret Results"],
            ["   - Check 'delta/fact' values (closer to 0 = better)"],
            ["   - Look at genetic potential (S1 cell)"],
            ["   - Compare with ML accuracy from your thesis"],
            [""],
            ["IMPORTANT NOTES:"],
            [""],
            ["1. Data Format:"],
            ["   - Values are RANKED (1, 2, 3, ...) not raw counts"],
            ["   - Direction matters (0 = higher is better, 1 = lower is better)"],
            ["   - Y values multiplied by 1000 (COCO convention)"],
            [""],
            ["2. Correlation Values:"],
            ["   - Positive correlation: attribute increases with positive sentiment"],
            ["   - Negative correlation: attribute increases with negative sentiment"],
            [""],
            ["3. Sample Size:"],
            [f"   - Currently using {n_samples} reviews"],
            ["   - You can increase this by changing n_samples parameter"],
            ["   - Recommended: 20-100 objects for COCO"],
            [""],
            ["4. Attributes Explained:"],
            ["   - positive_words: Count of positive sentiment words"],
            ["   - negative_words: Count of negative sentiment words"],
            ["   - exclamations: Number of ! marks (emotion intensity)"],
            ["   - review_length: Total words in review"],
            ["   - capital_ratio: Percentage of capital letters (shouting)"],
            ["   - questions: Number of ? marks (uncertainty)"],
            ["   - pos_neg_ratio: Ratio of positive to negative words"],
            ["   - word_diversity: Unique words / total words"],
        ]
        
        for row_idx, row_content in enumerate(instructions, start=1):
            ws_instructions.cell(row_idx, 1, row_content[0])
            if "Step" in row_content[0] or "IMPORTANT" in row_content[0] or "HOW TO" in row_content[0]:
                ws_instructions.cell(row_idx, 1).font = Font(bold=True, size=12)
        
        ws_instructions.column_dimensions['A'].width = 80
        
        # ==================================================================
        # SHEET 4: Correlation Summary
        # ==================================================================
        ws_corr = wb.create_sheet("Correlation Summary")
        
        ws_corr['A1'] = "Attribute"
        ws_corr['B1'] = "Correlation with Sentiment"
        ws_corr['C1'] = "Direction"
        ws_corr['D1'] = "Interpretation"
        
        for cell in ws_corr[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for idx, (attr, corr) in enumerate(sorted_attrs, start=2):
            ws_corr.cell(idx, 1, attr)
            ws_corr.cell(idx, 2, round(corr, 4))
            ws_corr.cell(idx, 3, "Positive" if corr > 0 else "Negative")
            
            if corr > 0:
                interpretation = "Higher values → More positive sentiment"
            else:
                interpretation = "Higher values → More negative sentiment"
            ws_corr.cell(idx, 4, interpretation)
        
        for col in ['A', 'B', 'C', 'D']:
            ws_corr.column_dimensions[col].width = 30
        
        # Save workbook
        wb.save(output_file)
        
        print(f"   [SUCCESS] Created: {output_file}")
        print(f"\n5. Summary:")
        print(f"   - Objects (reviews): {len(ranked_df)}")
        print(f"   - Attributes: {len(attribute_cols)}")
        print(f"   - Sheets created: 4")
        print(f"     1. OAM Matrix (raw data)")
        print(f"     2. COCO Input Format (ranked, ready to use)")
        print(f"     3. Instructions (how to use)")
        print(f"     4. Correlation Summary (interpretation)")
        
        print(f"\n{'='*70}")
        print("NEXT STEPS:")
        print(f"{'='*70}")
        print(f"1. Open {output_file}")
        print(f"2. Go to 'COCO Input Format' sheet")
        print(f"3. Copy data from row 6 onwards (columns B to last)")
        print(f"4. Visit: https://miau.my-x.hu/myx-free/coco/beker_std.php")
        print(f"5. Paste data and run analysis")
        print(f"6. Compare COCO results with your ML accuracy ({88.30}%)")
        
        return oam_df, ranked_df, correlations


# Run the export
if __name__ == "__main__":
    
    exporter = COCOExporter()
    
    # Export with different sample sizes
    print("\nChoose sample size for COCO analysis:")
    print("- Small (30 reviews): Fast, good for testing")
    print("- Medium (50 reviews): Balanced, recommended")
    print("- Large (100 reviews): More data, slower processing")
    
    # Default: 50 reviews
    n_samples = 50
    
    oam_df, ranked_df, correlations = exporter.create_oam_for_coco(
        n_samples=n_samples
    )
    
    print("\n" + "="*70)
    print("EXPORT COMPLETE!")
    print("="*70)
    print("\nFile 'OAM_for_COCO.xlsx' is ready for COCO analysis!")
